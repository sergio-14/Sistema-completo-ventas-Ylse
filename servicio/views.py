# views.py
from django.shortcuts import render,redirect,get_object_or_404
from django.views.generic import ListView, CreateView, DetailView
from .models import Venta,Producto,DetalleVenta
from .forms import VentaForm
from .forms import VentaForm, DetalleVentaForm
from django.forms import formset_factory
    
from django.contrib.auth import login, logout
from django.shortcuts import render, redirect
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Prefetch, Count, Q

from django.contrib.auth.models import Group
from django.contrib import messages
from .models import User, Cliente, Empleado,DivisionEmpleado,GastoDiario
from .forms import CustomUserCreationForm, CustomUserChangeForm

from .models import ProductoRetornable, Producto
from .forms import ProductoRetornableForm, ProductoForm
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required

def home(request):
    form = AuthenticationForm()
    return render(request, 'home.html', {'form': form})

def custom_login(request):
    if request.method == 'POST':
        form = AuthenticationForm(data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos. Inténtalo de nuevo.')
            return redirect('home')  # Asegúrate de que 'login' sea el nombre correcto de la URL de tu login
    else:
        form = AuthenticationForm()
    return render(request, 'home.html', {'form': form})

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from decimal import Decimal
from .models import Cliente, RegistroVenta, ProductosiRetornable
from .forms import CambiarEstadoForm  # Asegúrate de importar el formulario

@login_required
def dashboard(request):
    form = CambiarEstadoForm()  # Definir el formulario por defecto
    
    if request.method == 'POST':
        if 'registrar_pago' in request.POST:  # Verifica si se está registrando un pago
            cliente_id = request.POST.get('cliente_id')
            anticipo = Decimal(request.POST.get('anticipo', '0.0'))
            saldo_actual = Decimal(request.POST.get('saldo_actual', '0.0'))
            deuda_actual = Decimal(request.POST.get('deuda_actual', '0.0'))

            cliente = get_object_or_404(Cliente, id=cliente_id)

            nuevo_saldo = saldo_actual
            nueva_deuda = deuda_actual

            if anticipo <= saldo_actual:
                nuevo_saldo = saldo_actual - anticipo
            else:
                exceso = anticipo - saldo_actual
                nuevo_saldo = Decimal('0.0')
                nueva_deuda = max(deuda_actual - exceso, Decimal('0.0'))

            empleado = request.user.empleado

            RegistroVenta.objects.create(
                cliente=cliente,
                empleado=empleado,
                total=Decimal('0.0'),
                anticipo=anticipo,
                saldo_pendiente=nuevo_saldo,
                deuda_anterior=nueva_deuda
            )
            return redirect('venta_list')

        elif 'cambiar_estado' in request.POST:  # Verifica si se está cambiando el estado de productos
            form = CambiarEstadoForm(request.POST)
            if form.is_valid():
                productos_seleccionados = form.cleaned_data['productos']
                productos_seleccionados.update(estado="Disponible")  # Cambiar estado
                return redirect('dashboard')  # Redirigir al dashboard

    productos_ocupados = ProductosiRetornable.objects.filter(estado="Ocupado")  # Obtener productos ocupados
    clientes = Cliente.objects.all()  # Obtener clientes para el formulario

    return render(request, 'dashboard.html', {
        'user': request.user,
        'form': form,
        'productos_ocupados': productos_ocupados,
        'clientes': clientes,
    })


from .forms import CambiarEstadoForm

def cambiar_estado_productos(request):
    if request.method == "POST":
        form = CambiarEstadoForm(request.POST)
        if form.is_valid():
            productos_seleccionados = form.cleaned_data['productos']
            productos_seleccionados.update(estado="Disponible")  # Cambiar el estado
            return redirect('cambiar_estado')  # Redirigir a la misma página o donde prefieras
    else:
        form = CambiarEstadoForm()

    return render(request, "cambiar_estado.html", {"form": form})


from django.contrib.auth.decorators import login_required

from .forms import EmpleadoProductoForm


def lista_empleados(request):
    empleados = Empleado.objects.all()
    return render(request, 'empleados/lista_empleados.html', {'empleados': empleados})

def asignar_productos(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == 'POST':
        form = EmpleadoProductoForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            # Redirige a la vista de detalle o a la lista de empleados
            return redirect('lista_empleados')
    else:
        form = EmpleadoProductoForm(instance=empleado)
    
    return render(request, 'asignar_productos.html', {
        'form': form,
        'empleado': empleado,
    })

from .models import Cliente
from .forms import ClienteForm

def lista_clientes(request):
    clientes = Cliente.objects.all()
    return render(request, 'clientes/lista_clientes.html', {'clientes': clientes})


def editar_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    
    if request.method == 'POST':
        form = ClienteForm(request.POST, request.FILES, instance=cliente)
        if form.is_valid():
            form.save()
            return redirect('lista_clientes')
    else:
        form = ClienteForm(instance=cliente)
    
    return render(request, 'clientes/editar_cliente.html', {'form': form, 'cliente': cliente})


from .forms import EmpleadorutForm

def editar_empleado(request, empleado_id):
    empleado = get_object_or_404(Empleado, id=empleado_id)
    
    if request.method == 'POST':
        form = EmpleadorutForm(request.POST, instance=empleado)
        if form.is_valid():
            form.save()
            return redirect('lista_empleados')
    else:
        form = EmpleadorutForm(instance=empleado)
    
    return render(request, 'empleados/editar_empleado.html', {
        'form': form,
        'empleado': empleado,
    })

from django.shortcuts import render
from .models import Location, DivisionEmpleado, Empleado
from scipy.spatial.distance import euclidean
from itertools import permutations

from django.shortcuts import render
from .models import Location
from scipy.spatial.distance import euclidean
from itertools import permutations


def resolver_tsp(coordenadas, inicio):
    """
    Resuelve el problema del viajante (TSP) de forma ingenua generando todas las
    permutaciones de las ubicaciones (excluyendo el inicio) y retornando la secuencia
    con la menor distancia total. La ruta retorna al inicio.
    """
    ids = [loc[0] for loc in coordenadas]
    distancias = {
        (a[0], b[0]): euclidean((a[1], a[2]), (b[1], b[2]))
        for a in coordenadas for b in coordenadas if a != b
    }
    mejor_ruta = None
    mejor_distancia = float('inf')
    for perm in permutations(ids[1:]):
        ruta_actual = [inicio['id']] + list(perm) + [inicio['id']]
        distancia_actual = sum(
            distancias[(ruta_actual[i], ruta_actual[i+1])]
            for i in range(len(ruta_actual) - 1)
        )
        if distancia_actual < mejor_distancia:
            mejor_distancia = distancia_actual
            mejor_ruta = ruta_actual
    return mejor_ruta

def mapas(request):
    """
    Vista para mostrar ubicaciones y, opcionalmente, calcular la ruta óptima.
    Si se pasa el parámetro 'ruta_id', se filtran las ubicaciones de esa ruta.
    Si además se pasa 'mostrar_ruta=1', se calcula la ruta óptima (TSP) y se incluye.
    """
    # Obtener todas las rutas disponibles (DivisionEmpleado)
    rutas = DivisionEmpleado.objects.all()

    # Verificar si se ha pasado un filtro de ruta
    ruta_id = request.GET.get("ruta_id")
    if ruta_id:
        locations = list(
            Location.objects.filter(ruta__id=ruta_id)
            .select_related('cliente', 'cliente__usuario', 'ruta')
            .values(
                'id',
                'latitude',
                'longitude',
                'cliente__negocio',
                'cliente__imagen',
                'cliente__usuario__nombre',
                'cliente__usuario__apellido',
                'ruta__ruta'  # Nombre de la ruta
            )
        )
    else:
        # Si no se pasa ruta_id, se muestran TODAS las ubicaciones
        locations = list(
            Location.objects.all()
            .select_related('cliente', 'cliente__usuario', 'ruta')
            .values(
                'id',
                'latitude',
                'longitude',
                'cliente__negocio',
                'cliente__imagen',
                'cliente__usuario__nombre',
                'cliente__usuario__apellido',
                'ruta__ruta'
            )
        )

    ruta_optima = None
    # Solo se calcula la ruta óptima si se filtra por una ruta y se indica mostrar la ruta
    if ruta_id and request.GET.get("mostrar_ruta") == "1" and locations:
        inicio = locations[0]
        coordenadas = [(loc['id'], loc['latitude'], loc['longitude']) for loc in locations]
        ruta_ids = resolver_tsp(coordenadas, inicio)
        ruta_optima = [next(loc for loc in locations if loc['id'] == loc_id) for loc_id in ruta_ids]

    context = {
        'locations': locations,
        'ruta_optima': ruta_optima,
        'rutas': rutas,  # Pasamos las rutas disponibles al template
    }
    return render(request, 'mapas.html', context)



def buscar_cliente(request):
    negocio = request.GET.get('negocio')
    if not negocio:
        return JsonResponse({'error': 'Debe ingresar el nombre del negocio'}, status=400)
    
    cliente = Cliente.objects.filter(negocio__icontains=negocio).first()
    if not cliente:
        return JsonResponse({'error': 'Cliente no encontrado'}, status=404)
    
    # Obtener el último registro de venta para obtener el estado actual
    ultimo_registro = RegistroVenta.objects.filter(cliente=cliente).first()
    if ultimo_registro:
        saldo_pendiente = ultimo_registro.saldo_pendiente
        deuda_anterior = ultimo_registro.deuda_anterior
    else:
        # En caso de que no exista registro, se pueden iniciar en 0 o con otros valores base.
        saldo_pendiente = Decimal('0.0')
        deuda_anterior = Decimal('0.0')
    
    data = {
        'cliente_id': cliente.id,
        'negocio': cliente.negocio,
        'saldo_pendiente': str(saldo_pendiente),
        'deuda_anterior': str(deuda_anterior),
    }
    return JsonResponse(data)

def registrar_pago(request):
    if request.method == 'POST':
        cliente_id = request.POST.get('cliente_id')
        anticipo = Decimal(request.POST.get('anticipo', '0.0'))
        # Opcional: capturar los saldos actuales enviados desde el formulario (si se confía en ellos)
        saldo_actual = Decimal(request.POST.get('saldo_actual', '0.0'))
        deuda_actual = Decimal(request.POST.get('deuda_actual', '0.0'))
        
        cliente = get_object_or_404(Cliente, id=cliente_id)
        
        # Lógica de actualización:
        # Si el anticipo es menor o igual al saldo pendiente, se descuenta solo de éste.
        # Si el anticipo excede el saldo pendiente, se descuenta el exceso de la deuda anterior.
        nuevo_saldo = saldo_actual
        nueva_deuda = deuda_actual
        
        if anticipo <= saldo_actual:
            nuevo_saldo = saldo_actual - anticipo
        else:
            exceso = anticipo - saldo_actual
            nuevo_saldo = Decimal('0.0')
            if exceso <= deuda_actual:
                nueva_deuda = deuda_actual - exceso
            else:
                nueva_deuda = Decimal('0.0')
        
        # Se asume que el empleado se obtiene del usuario logueado.
        # Ajusta esta parte según cómo manejes la autenticación y relación con Empleado.
        empleado = request.user.empleado
        
        # Registrar la operación como una venta con total = 0
        RegistroVenta.objects.create(
            cliente=cliente,
            empleado=empleado,
            total=Decimal('0.0'),
            anticipo=anticipo,
            saldo_pendiente=nuevo_saldo,
            deuda_anterior=nueva_deuda
        )
        return redirect('lista_ventas')  # Redirige a la lista de ventas o a la confirmación que prefieras.
    else:
        return render(request, 'buscador_cliente.html')





from django.utils import timezone
from datetime import datetime
import pytz

# Listar ventas
def venta_list(request):
    # Obtener la fecha actual en la zona horaria local
    la_paz_tz = pytz.timezone('America/La_Paz')
    hoy = timezone.now().astimezone(la_paz_tz).date()

    # Filtrar ventas y gastos solo del día actual
    ventas = RegistroVenta.objects.filter(fecha__date=hoy).order_by('-fecha') 
    gastos = GastoDiario.objects.filter(fecha=hoy).order_by('-fecha') 

    # Cálculos de totales
    total_general = sum(venta.total for venta in ventas)
    total_anticipo = sum(venta.anticipo for venta in ventas)
    monto_total_gastos = sum(gasto.monto for gasto in gastos)
    total_ganancias = total_anticipo - monto_total_gastos

    context = {
        'ventas': ventas,
        'gastos': gastos,
        'total_general': total_general,
        'monto_total_gastos': monto_total_gastos,
        'total_anticipo': total_anticipo,
        'total_ganancias': total_ganancias,
    }

    return render(request, 'ventas/venta_list.html', context)

from django.db.models import Sum

def venta_listrepo(request):
    # Definir la zona horaria de Bolivia
    la_paz_tz = pytz.timezone("America/La_Paz")

    # Obtener parámetros de filtrado desde GET
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    ruta = request.GET.get("ruta")

    # Convertir fechas a objetos datetime con zona horaria
    if start_date and end_date:
        try:
            start_date_obj = la_paz_tz.localize(datetime.strptime(start_date, "%Y-%m-%d")).replace(hour=0, minute=0, second=0)
            end_date_obj = la_paz_tz.localize(datetime.strptime(end_date, "%Y-%m-%d")).replace(hour=23, minute=59, second=59)
        except ValueError:
            start_date_obj = end_date_obj = None
    else:
        start_date_obj = end_date_obj = None

    # Obtener la fecha actual con zona horaria
    hoy = timezone.now().astimezone(la_paz_tz).replace(hour=0, minute=0, second=0)

    # Obtener todos los registros
    ventas = RegistroVenta.objects.all().order_by("-fecha")
    gastos = GastoDiario.objects.all().order_by("-fecha")

    # Filtrar por rango de fechas (o por hoy si no se proporcionan)
    if start_date_obj and end_date_obj:
        ventas = ventas.filter(fecha__range=(start_date_obj, end_date_obj))
        gastos = gastos.filter(fecha__range=(start_date_obj, end_date_obj))
    else:
        ventas = ventas.filter(fecha__gte=hoy, fecha__lt=hoy.replace(hour=23, minute=59, second=59))
        gastos = gastos.filter(fecha__gte=hoy, fecha__lt=hoy.replace(hour=23, minute=59, second=59))

    # Filtrar por ruta del empleado si se especifica
    if ruta:
        ventas = ventas.filter(empleado__division__ruta=ruta)
        gastos = gastos.filter(empleado__division__ruta=ruta)

    # Cálculos de totales optimizados con aggregate
    total_general = ventas.aggregate(Sum("total"))["total__sum"] or 0
    total_anticipo = ventas.aggregate(Sum("anticipo"))["anticipo__sum"] or 0
    monto_total_gastos = gastos.aggregate(Sum("monto"))["monto__sum"] or 0
    total_ganancias = total_anticipo - monto_total_gastos

    # Obtener todas las rutas disponibles para el filtro
    rutas = DivisionEmpleado.objects.all()

    # Paginación: 10 elementos por página para ventas y gastos
    ventas_paginator = Paginator(ventas, 10)
    gastos_paginator = Paginator(gastos, 10)

    page_number_ventas = request.GET.get("page_ventas")
    page_number_gastos = request.GET.get("page_gastos")

    ventas_page = ventas_paginator.get_page(page_number_ventas)
    gastos_page = gastos_paginator.get_page(page_number_gastos)
    
    

    # Si hay un error (ejemplo: rango de fechas incorrecto)
    if start_date and end_date and start_date > end_date:
        messages.error(request, "La fecha de inicio no puede ser mayor que la fecha final.")

    context = {
        "ventas": ventas_page,
        "gastos": gastos_page,
        "total_general": total_general,
        "total_anticipo": total_anticipo,
        "monto_total_gastos": monto_total_gastos,
        "total_ganancias": total_ganancias,
        # Datos de filtros para mantener los valores seleccionados en el formulario
        "start_date": start_date if start_date else "",
        "end_date": end_date if end_date else "",
        "ruta_selected": ruta,
        "rutas": rutas,
    }

    return render(request, "ventas/venta_listrepo.html", context)


import openpyxl
from openpyxl.styles import Font
from django.http import HttpResponse
from datetime import datetime
import pytz

def exportar_excel(request):
    # Definir la zona horaria de Bolivia
    la_paz_tz = pytz.timezone("America/La_Paz")

    # Obtener parámetros de filtrado desde GET
    start_date = request.GET.get("start_date")
    end_date = request.GET.get("end_date")
    ruta = request.GET.get("ruta")

    # Convertir fechas a objetos datetime con zona horaria
    if start_date and end_date:
        try:
            start_date_obj = la_paz_tz.localize(datetime.strptime(start_date, "%Y-%m-%d")).replace(hour=0, minute=0, second=0)
            end_date_obj = la_paz_tz.localize(datetime.strptime(end_date, "%Y-%m-%d")).replace(hour=23, minute=59, second=59)
        except ValueError:
            start_date_obj = end_date_obj = None
    else:
        start_date_obj = end_date_obj = None

    # Obtener todos los registros
    ventas = RegistroVenta.objects.all().order_by("-fecha")
    gastos = GastoDiario.objects.all().order_by("-fecha")

    # Aplicar filtros
    if start_date_obj and end_date_obj:
        ventas = ventas.filter(fecha__range=(start_date_obj, end_date_obj))
        gastos = gastos.filter(fecha__range=(start_date_obj, end_date_obj))
    if ruta:
        ventas = ventas.filter(empleado__division__ruta=ruta)
        gastos = gastos.filter(empleado__division__ruta=ruta)

    # Crear el libro de Excel y la hoja
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Ventas y Gastos"

    # Estilo para encabezado
    bold_font = Font(bold=True)

    # **Encabezado para la tabla de Ventas**
    ws.append(["LISTA DE VENTAS"])  # Título
    ws.append(["Fecha", "Cliente", "Empleado", "Total", "Anticipo", "Saldo"])  # Encabezado

    for cell in ws[2]:
        cell.font = bold_font

    # Agregar datos de ventas
    for venta in ventas:
        saldo = venta.total - venta.anticipo
        ws.append([
            venta.fecha.strftime("%Y-%m-%d"),
            str(venta.cliente),
            str(venta.empleado),
            venta.total,
            venta.anticipo,
            saldo
        ])

    # Línea vacía entre tablas
    ws.append([])
    ws.append(["LISTA DE GASTOS"])
    ws.append(["Fecha", "Descripción", "Monto"])  # Encabezado

    for cell in ws[ws.max_row]:
        cell.font = bold_font

    # Agregar datos de gastos
    for gasto in gastos:
        ws.append([
            gasto.fecha.strftime("%Y-%m-%d"),
            gasto.descripcion,
            gasto.monto
        ])

    # Totales
    total_general = ventas.aggregate(Sum("total"))["total__sum"] or 0
    total_anticipo = ventas.aggregate(Sum("anticipo"))["anticipo__sum"] or 0
    monto_total_gastos = gastos.aggregate(Sum("monto"))["monto__sum"] or 0
    total_ganancias = total_anticipo - monto_total_gastos

    # Línea vacía para separar totales
    ws.append([])
    ws.append(["Resumen de Totales"])
    ws.append(["Total Ventas", total_general])
    ws.append(["Total Anticipos", total_anticipo])
    ws.append(["Total Gastos", monto_total_gastos])
    ws.append(["Total Ganancias", total_ganancias])

    # Ajustar el ancho de las columnas
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter  # Obtener la letra de la columna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Preparar la respuesta HTTP para descargar el archivo
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="Reporte_Ventas_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx"'
    
    wb.save(response)
    return response


from django.db.models import OuterRef, Subquery
from django.shortcuts import render
from .models import RegistroVenta, Cliente

def ultima_venta_clientes(request):
    # Subconsulta para obtener la fecha del ultimo registro de cada cliente
    ultimas_ventas = RegistroVenta.objects.filter(
        cliente=OuterRef('pk')
    ).order_by('-fecha')

    # Anotamos el ultimo registro de venta en cada cliente
    clientes_con_ultima_venta = Cliente.objects.annotate(
        ultima_venta_id=Subquery(ultimas_ventas.values('id')[:1]),
        ultima_venta_fecha=Subquery(ultimas_ventas.values('fecha')[:1]),
        ultima_venta_saldo=Subquery(ultimas_ventas.values('saldo_pendiente')[:1]),
        ultima_venta_total=Subquery(ultimas_ventas.values('total')[:1]),
    ).filter(ultima_venta_id__isnull=False).order_by('-ultima_venta_fecha')

    context = {
        'clientes': clientes_con_ultima_venta,
    }
    return render(request, 'ventas/ultima_venta_clientes.html', context)


from django.db.models import F, Window
from django.db.models.functions import RowNumber
from .models import RegistroDetalleVenta

def clientes_ocupados(request):
    # Filtrar detalles con producto retornable en estado "Ocupado"
    # y usar Window para asignar un número de fila por cada producto, ordenando por la fecha de venta de forma descendente.
    detalles_ocupados = RegistroDetalleVenta.objects.filter(
        producto_retornable__estado='Ocupado'
    ).annotate(
        rn=Window(
            expression=RowNumber(),
            partition_by=[F('producto_retornable')],
            order_by=F('venta__fecha').desc()
        )
    ).filter(rn=1).select_related('venta__cliente', 'producto_retornable')
    
    clientes_productos = {}
    for detalle in detalles_ocupados:
        cliente = detalle.venta.cliente
        producto = detalle.producto_retornable
        # Agrupamos usando la PK del cliente
        if cliente.pk not in clientes_productos:
            clientes_productos[cliente.pk] = {
                'cliente': cliente,
                'productos': []
            }
        clientes_productos[cliente.pk]['productos'].append({
            'descripcion': producto.descripcion,
            'estado': producto.estado,
            'tipo': producto.tipo_producto.nombre,
        })

    context = {'clientes_productos': clientes_productos.values()}
    return render(request, 'ventas/clientes_ocupados.html', context)

# seccion usuarios 
def agregar_usuario(request):
    if request.method == 'POST':
        form = CustomUserCreationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            
            # Asignar grupos
            grupos = form.cleaned_data.get('groups')
            for grupo in grupos:
                user.groups.add(grupo)
            
            # Verificar si es Cliente o Empleado
            tipo = request.POST.get('tipo_usuario')
            if tipo == 'cliente':
                Cliente.objects.create(
                    usuario=user, 
                    negocio=request.POST.get('negocio'), 
                    contacto=request.POST.get('contacto'),
                    imagen=request.FILES.get('imagen')
                )
            elif tipo == 'empleado':
                division_id = request.POST.get('division')
                Empleado.objects.create(
                    usuario=user, 
                    domicilio=request.POST.get('domicilio'), 
                    division_id=division_id
                )
            
            messages.success(request, "Usuario creado exitosamente.")
            return redirect('listar_usuarios')
    else:
        form = CustomUserCreationForm()
    
    # Obtener las divisiones disponibles
    divisiones = DivisionEmpleado.objects.all()
    
    return render(request, 'usuarios/agregar_usuario.html', {'form': form, 'divisiones': divisiones})




def editar_usuario(request, user_id):
    user = get_object_or_404(User, id=user_id)
    if request.method == 'POST':
        form = CustomUserChangeForm(request.POST, request.FILES, instance=user)
        if form.is_valid():
            form.save()
            messages.success(request, "Usuario actualizado exitosamente.")
            return redirect('listar_usuarios')
        else:
            # Agregar un mensaje de error si el formulario no es válido
            messages.error(request, "Hubo un error con el formulario.")
            print(form.errors)  # Esto te permitirá ver los errores del formulario en la consola
    else:
        form = CustomUserChangeForm(instance=user)
    return render(request, 'usuarios/editar_usuario.html', {'form': form, 'user': user})



from django.contrib.auth.decorators import login_required, user_passes_test

@login_required
def toggle_user_status(request, user_id):
    user = get_object_or_404(User, id=user_id)
    
    if user.is_active:
        user.is_active = False
        messages.success(request, f'El usuario {user.email} ha sido deshabilitado.')
    else:
        user.is_active = True
        messages.success(request, f'El usuario {user.email} ha sido habilitado.')

    user.save()
    return redirect('listar_usuarios') 

from django.db.models import Q

def listar_usuarios(request):
    query = request.GET.get('q', '') 
    usuarios = User.objects.all()

    if query:
        usuarios = usuarios.filter(
            Q(nombre__icontains=query) | 
            Q(apellido__icontains=query) | 
            Q(apellidoM__icontains=query)
        )

    # Paginación (5 usuarios por página)
    paginator = Paginator(usuarios, 8)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    return render(request, 'usuarios/listar_usuarios.html', {
        'page_obj': page_obj,
        'query': query
    })

"""def listar_usuarios(request):
    usuarios = User.objects.all()
    return render(request, 'usuarios/listar_usuarios.html', {'usuarios': usuarios})
"""

from decimal import Decimal
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from .forms import VentaForm

def registrar_venta(request):
    empleado = request.user.empleado  # Obtén el empleado relacionado con el usuario
    if request.method == 'POST':
        form = VentaForm(request.POST)
        if form.is_valid():
            venta = form.save(commit=False)
            venta.empleado = empleado  # Asigna el empleado automáticamente
            venta.save()  # Guarda la venta
            form.save_m2m()  # Guarda la relación Many-to-Many con productos
            
            # Actualiza el estado de los productos retornables
            for producto in venta.productos.all():
                producto.estado = 'Ocupado'
                producto.save()
            
            return redirect('venta_list')  # Redirige a la lista de ventas
    else:
        form = VentaForm()

    return render(request, 'registrar_venta.html', {'form': form})


@login_required
def obtener_saldo_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    ultima_venta = RegistroVenta.objects.filter(cliente=cliente).order_by('-fecha').first()
    saldo_pendiente = float(ultima_venta.saldo_pendiente) if ultima_venta else 0
    return JsonResponse({'saldo_pendiente': saldo_pendiente})



from decimal import Decimal

def crear_venta(request):
    DetalleVentaFormSet = formset_factory(DetalleVentaForm, extra=0)
    empleado = request.user.empleado  # Obtén el empleado relacionado con el usuario

    if request.method == 'POST':
        
        form = VentaForm(request.POST)
        formset = DetalleVentaFormSet(request.POST)

        if form.is_valid() and formset.is_valid():
            # Guardar la venta
            venta = form.save(commit=False)
            venta.empleado = empleado  # Asigna el empleado automáticamente
            venta.save()  # Guarda la venta primero
            form.save_m2m()  # Guarda la relación Many-to-Many con productos

            # Actualizar estado de los productos retornables
            for producto in venta.productos.all():
                producto.estado = 'Ocupado'
                producto.save()

            # Guardar el detalle de la venta
            for detalle_form in formset:
                if detalle_form.cleaned_data:
                    producto = detalle_form.cleaned_data['producto']
                    cantidad = detalle_form.cleaned_data['cantidad']
                    subtotal = cantidad * producto.precio  # Precio del producto desde el modelo

                    # Crear y guardar el detalle de la venta
                    DetalleVenta.objects.create(
                        venta=venta,
                        producto=producto,
                        cantidad=cantidad,
                        subtotal=subtotal
                    )

            # Redirigir a la lista de ventas
            return redirect('venta_list')
        else:
            # Depuración de errores
            if not form.is_valid():
                print("Errores en el formulario principal:")
                print(form.errors)

            if not formset.is_valid():
                print("Errores en el formset:")
                for detalle_form in formset:
                    print(detalle_form.errors)
    else:
        form = VentaForm()
        formset = DetalleVentaFormSet()

    productos = Producto.objects.values('id', 'tipo_producto__nombre', 'precio')
    return render(request, 'ventas/venta_form.html', {
        'form': form,
        'formset': formset,
        'productos': productos,
    })



from django.http import JsonResponse
from .models import Cliente


    
    



# Ver detalles de venta
class VentaDetailView(DetailView):
    model = Venta
    model = GastoDiario
    template_name = 'ventas/venta_detail.html'
    context_object_name = 'venta, gasto_diario'
    
def detalle_venta(request, venta_id):
    # Obtiene la venta específica o lanza un error 404 si no existe
    venta = get_object_or_404(RegistroVenta, id=venta_id)

    # Obtiene los detalles asociados a la venta, cargando el tipo de producto según corresponda
    detalles = venta.detalles.select_related('producto_no_retornable__tipo_producto', 'producto_retornable__tipo_producto')

    # Pasamos la venta y sus detalles al contexto
    return render(request, 'detalle_venta.html', {
        'venta': venta,
        'detalles': detalles,
    })

from django.http import HttpResponse
from .utils import render_to_pdf

def detalle_venta_pdf(request, venta_id):
    venta = get_object_or_404(RegistroVenta, id=venta_id)
    detalles = venta.detalles.select_related('producto_no_retornable__tipo_producto', 'producto_retornable__tipo_producto')
    
    context = {
        'venta': venta,
        'detalles': detalles,
    }
    
    pdf = render_to_pdf('detalle_venta_pdf.html', context)
    if pdf:
        response = HttpResponse(pdf, content_type='application/pdf')
        filename = f"DetalleVenta.pdf"
        content = f"attachment; filename={filename}"
        response['Content-Disposition'] = content
        return response
    return HttpResponse("No se pudo generar el PDF", status=400)    


def custom_logout(request):
    logout(request)
    return redirect('home')


from .models import ProductosiRetornable, ProductonoRetornable
from .forms import ProductosiRetornableForm, ProductonoRetornableForm
#productos
# Listar Productos
def listar_productos(request):
    if request.method == "POST":
        form_type = request.POST.get("form_type")
        
        # Procesar para Producto Retornable
        if form_type == "producto_retornable":
            producto_retornable_id = request.POST.get("producto_retornable_id")
            if producto_retornable_id:  # Editar
                producto = get_object_or_404(ProductosiRetornable, id=producto_retornable_id)
                form_retornable = ProductosiRetornableForm(request.POST, instance=producto)
            else:  # Crear
                form_retornable = ProductosiRetornableForm(request.POST)
            
            if form_retornable.is_valid():
                form_retornable.save()
                return redirect("listar_productos")
        
        # Procesar para Producto No Retornable
        elif form_type == "producto_no_retornable":
            producto_no_retornable_id = request.POST.get("producto_no_retornable_id")
            if producto_no_retornable_id:  # Editar
                producto = get_object_or_404(ProductonoRetornable, id=producto_no_retornable_id)
                form_no_retornable = ProductonoRetornableForm(request.POST, instance=producto)
            else:  # Crear
                form_no_retornable = ProductonoRetornableForm(request.POST)
            
            if form_no_retornable.is_valid():
                form_no_retornable.save()
                return redirect("listar_productos")
    
    else:
        form_retornable = ProductosiRetornableForm()
        form_no_retornable = ProductonoRetornableForm()
    
    # Listado y paginación
    productos_retornables_list = ProductosiRetornable.objects.all()
    productos_list = ProductonoRetornable.objects.all()
    
    paginator_retornables = Paginator(productos_retornables_list, 5)
    paginator_productos = Paginator(productos_list, 5)
    
    page_number_retornables = request.GET.get('page_retornables')
    page_number_productos = request.GET.get('page_productos')
    
    page_retornables = paginator_retornables.get_page(page_number_retornables)
    page_productos = paginator_productos.get_page(page_number_productos)
    
    context = {
        'page_retornables': page_retornables,
        'page_productos': page_productos,
        'form_retornable': form_retornable,
        'form_no_retornable': form_no_retornable,
    }
    return render(request, 'productos/listar_productos.html', context)
    
# Crear ProductoRetornable
def crear_producto_retornable(request):
    if request.method == 'POST':
        form = ProductoRetornableForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoRetornableForm()
    return render(request, 'productos/formulario_re.html', {'form': form})

# Editar ProductoRetornable
def editar_producto_retornable(request, pk):
    producto = get_object_or_404(ProductosiRetornable, pk=pk)
    if request.method == 'POST':
        form = ProductoRetornableForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoRetornableForm(instance=producto)
    return render(request, 'productos/formulario_re.html', {'form': form})

# Crear Producto
def crear_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm()
    return render(request, 'productos/formulario_de.html', {'form': form})

# Editar Producto
def editar_producto(request, pk):
    producto = get_object_or_404(ProductonoRetornable, pk=pk)
    if request.method == 'POST':
        form = ProductoForm(request.POST, instance=producto)
        if form.is_valid():
            form.save()
            return redirect('listar_productos')
    else:
        form = ProductoForm(instance=producto)
    return render(request, 'productos/formulario_de.html', {'form': form})



from django.views.generic import ListView, CreateView, UpdateView
from django.urls import reverse_lazy
from .models import TipoProducto, Location, DivisionEmpleado, GastoDiario
from .forms import TipoProductoForm, LocationForm, DivisionEmpleadoForm, GastoDiarioForm

def combined_list(request):
    location_id = request.POST.get("location_id")  # Capturar el ID si es una edición

    if request.method == "POST":
        if location_id:  # Si hay un ID, estamos editando
            location = get_object_or_404(Location, id=location_id)
            form = LocationForm(request.POST, instance=location)
        else:  # Si no hay ID, estamos creando
            form = LocationForm(request.POST)

        if form.is_valid():
            form.save()
            return redirect("combined_list")  # Redirigir tras éxito

    else:
        form = LocationForm()

    # Paginación
    location_list = Location.objects.all().order_by('-id')  
    location_paginator = Paginator(location_list, 8)  
    location_page = request.GET.get('location_page')
    locations = location_paginator.get_page(location_page)

    context = {
        'locations': locations,
        'form': form
    }
    return render(request, 'funciones/combined_list.html', context)



def combined_lista(request):
    if request.method == "POST":
        form_type = request.POST.get("form_type")

        # 👉 Editar o crear Tipo Producto
        if form_type == "tipo_producto":
            tipo_producto_id = request.POST.get("tipo_producto_id")

            if tipo_producto_id:  # Editar
                tipo_producto = get_object_or_404(TipoProducto, id=tipo_producto_id)
                form = TipoProductoForm(request.POST, instance=tipo_producto)
            else:  # Crear
                form = TipoProductoForm(request.POST)

            if form.is_valid():
                form.save()
                return redirect("combined_lista")

        # 👉 Editar o crear Ruta
        elif form_type == "ruta":
            ruta_id = request.POST.get("ruta_id")

            if ruta_id:  # Editar
                ruta = get_object_or_404(DivisionEmpleado, id=ruta_id)
                formr = DivisionEmpleadoForm(request.POST, instance=ruta)
            else:  # Crear
                formr = DivisionEmpleadoForm(request.POST)

            if formr.is_valid():
                formr.save()
                return redirect("combined_lista")

    else:
        form = TipoProductoForm()
        formr = DivisionEmpleadoForm()

    # Paginación
    tipo_producto_list = TipoProducto.objects.all()
    tipo_producto_paginator = Paginator(tipo_producto_list, 4)
    tipo_producto_page = request.GET.get('tipo_producto_page')
    tipo_producto = tipo_producto_paginator.get_page(tipo_producto_page)

    division_empleado_list = DivisionEmpleado.objects.all()
    division_empleado_paginator = Paginator(division_empleado_list, 4)
    division_empleado_page = request.GET.get('division_empleado_page')
    division_empleados = division_empleado_paginator.get_page(division_empleado_page)

    context = {
        'form': form,
        'formr': formr,
        'tipo_producto': tipo_producto,
        'division_empleados': division_empleados,
    }
    return render(request, 'funciones/combined_lista.html', context)


class TipoProductoCreateView(CreateView):
    model = TipoProducto
    form_class = TipoProductoForm
    template_name = 'funciones/tipo_producto_form.html'
    success_url = reverse_lazy('combined_list')

class TipoProductoUpdateView(UpdateView):
    model = TipoProducto
    form_class = TipoProductoForm
    template_name = 'funciones/tipo_producto_form.html'
    success_url = reverse_lazy('combined_list')



class LocationCreateView(CreateView):
    model = Location
    form_class = LocationForm
    template_name = 'funciones/location_form.html'
    success_url = reverse_lazy('combined_list')
    


class LocationUpdateView(UpdateView):
    model = Location
    form_class = LocationForm
    template_name = 'funciones/location_form.html'
    success_url = reverse_lazy('combined_list')



class DivisionEmpleadoCreateView(CreateView):
    model = DivisionEmpleado
    form_class = DivisionEmpleadoForm
    template_name = 'funciones/division_empleado_form.html'
    success_url = reverse_lazy('combined_list_view')

class DivisionEmpleadoUpdateView(UpdateView):
    model = DivisionEmpleado
    form_class = DivisionEmpleadoForm
    template_name = 'funciones/division_empleado_form.html'
    success_url = reverse_lazy('combined_list_view')
    
from django.contrib.auth.mixins import LoginRequiredMixin

class GastoDiarioCreateView(LoginRequiredMixin, CreateView):
    model = GastoDiario
    form_class = GastoDiarioForm
    template_name = 'gastos/gasto_diario_form.html'
    success_url = reverse_lazy('venta_list')

    def form_valid(self, form):
        # Obtener el empleado relacionado con el usuario autenticado
        try:
            empleado = self.request.user.empleado  # Accede al modelo Empleado desde User
        except Empleado.DoesNotExist:
            # Si no existe el empleado, muestra un error
            form.add_error(None, "No tienes un perfil de empleado asignado.")
            return self.form_invalid(form)
        
        form.instance.empleado = empleado  # Asigna el empleado al formulario
        return super().form_valid(form)

class GastoDiarioUpdateView(UpdateView):
    model = GastoDiario
    form_class = GastoDiarioForm
    template_name = 'gastos/gasto_diario_form.html'
    success_url = reverse_lazy('venta_list')
    




from django.shortcuts import render, redirect
from django.views import View
from decimal import Decimal
from .forms import RegistroVentaForm
from .models import Cliente, Empleado, ProductonoRetornable, ProductosiRetornable, RegistroVenta, RegistroDetalleVenta

from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from decimal import Decimal

from decimal import Decimal
from django.shortcuts import render, redirect, get_object_or_404
from django.views import View
from django.contrib.auth.decorators import login_required
from django.utils.decorators import method_decorator
from .forms import RegistroVentaForm
from .models import RegistroVenta, RegistroDetalleVenta, ProductonoRetornable, ProductosiRetornable, Empleado

from django.shortcuts import render, redirect
from django.utils.decorators import method_decorator
from django.views import View
from django.contrib.auth.decorators import login_required
from django.core.exceptions import ObjectDoesNotExist
from .models import RegistroVenta, Empleado, ProductonoRetornable, ProductosiRetornable, RegistroDetalleVenta
from .forms import RegistroVentaForm

@method_decorator(login_required, name='dispatch')
class RegistrarVentaView(View):
    def get(self, request, *args, **kwargs):
        form = RegistroVentaForm()

        # Solo mostrar productos retornables disponibles (estado "Disponible")
        productos_no_retornables = ProductonoRetornable.objects.all()
        productos_retornables = ProductosiRetornable.objects.filter(estado='Disponible')

        return render(request, 'ventas/registrar_venta.html', {
            'form': form,
            'productos_no_retornables': productos_no_retornables,
            'productos_retornables': productos_retornables
        })

    def post(self, request, *args, **kwargs):
        form = RegistroVentaForm(request.POST)
        if form.is_valid():
            cliente = form.cleaned_data['cliente']
            anticipo = form.cleaned_data['anticipo']
            total_value = form.cleaned_data['total']
            saldo_value = form.cleaned_data['saldo_pendiente']
            deuda_value = form.cleaned_data['deuda_anterior']

            try:
                empleado = Empleado.objects.get(usuario=request.user)
            except ObjectDoesNotExist:
                return render(request, 'ventas/registrar_venta.html', {
                    'form': form,
                    'error': 'Empleado no encontrado para el usuario actual.'
                })

            productos_no_retornables = []
            cantidades_no_retornables = []
            productos_retornables = []
            cantidades_retornables = []

            # Procesar productos no retornables seleccionados
            for producto in ProductonoRetornable.objects.all():
                if form.cleaned_data.get(f'producto_no_retornable_{producto.id}'):
                    cantidad = form.cleaned_data.get(f'cantidad_no_retornable_{producto.id}', 0)
                    if cantidad > 0:
                        productos_no_retornables.append(producto)
                        cantidades_no_retornables.append(cantidad)

            # Procesar productos retornables seleccionados (solo los disponibles)
            for producto in ProductosiRetornable.objects.filter(estado='Disponible'):
                if form.cleaned_data.get(f'producto_retornable_{producto.id}'):
                    cantidad = form.cleaned_data.get(f'cantidad_retornable_{producto.id}', 0)
                    if cantidad > 0:
                        productos_retornables.append(producto)
                        cantidades_retornables.append(cantidad)

            if not productos_no_retornables and not productos_retornables:
                return render(request, 'ventas/registrar_venta.html', {
                    'form': form,
                    'error': 'Debes agregar al menos un producto con cantidad válida.'
                })

            # Crear la venta
            venta = RegistroVenta.objects.create(
                cliente=cliente,
                empleado=empleado,
                anticipo=anticipo,
                total=total_value,
                saldo_pendiente=saldo_value,
                deuda_anterior=deuda_value
            )

            # Registrar los detalles de la venta y ajustar stocks y estados
            for i, producto in enumerate(productos_no_retornables):
                cantidad = cantidades_no_retornables[i]
                subtotal = producto.precio * cantidad
                RegistroDetalleVenta.objects.create(
                    venta=venta,
                    producto_no_retornable=producto,
                    cantidad=cantidad,
                    precio=producto.precio,
                    subtotal=subtotal
                )

            for i, producto in enumerate(productos_retornables):
                cantidad = cantidades_retornables[i]
                subtotal = producto.precio * cantidad

                # Cambiar estado del producto retornable a "Ocupado"
                producto.estado = 'Ocupado'
                producto.save()

                RegistroDetalleVenta.objects.create(
                    venta=venta,
                    producto_retornable=producto,
                    cantidad=cantidad,
                    precio=producto.precio,
                    subtotal=subtotal
                )

            return redirect('venta_list')

        return render(request, 'ventas/registrar_venta.html', {'form': form})



class DetalleVentaView(View):
    def get(self, request, pk, *args, **kwargs):
        venta = RegistroVenta.objects.get(id=pk)
        detalles = venta.detalles.all()
        context = {
            'venta': venta,
            'detalles': detalles,
        }
        return render(request, 'detalle_venta.html', context)
    
    
    
    
    
from .models import RegistroVenta, RegistroDetalleVenta, ProductosiRetornable, ProductonoRetornable
def lista_ventas(request):
    ventas = RegistroVenta.objects.all()  # Recuperar todas las ventas

    # Puedes usar select_related para optimizar las consultas si es necesario
    # ventas = RegistroVenta.objects.select_related('cliente', 'empleado').all()

    return render(request, 'ventas/lista_ventas.html', {
        'ventas': ventas
    })

from django.shortcuts import get_object_or_404

def detalle_venta(request, venta_id):
    venta = get_object_or_404(RegistroVenta, id=venta_id)
    detalles = venta.detalles.all()  # Obtener los detalles de esta venta

    return render(request, 'ventas/detalle_venta.html', {
        'venta': venta,
        'detalles': detalles
    })